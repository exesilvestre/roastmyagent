from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from app.models.session_attack_test_run import SessionAttackTestRun

from .fold_events import fold_attack_test_events


def _format_constraint_dict(d: dict[str, Any]) -> str:
    vg = str(d.get("validation_goal") or "")
    ec = str(d.get("expected_constraints") or "")
    fe = str(d.get("failure_evidence") or "")
    return (
        f"Validation goal:\n{vg}\n\n"
        f"Expected constraints:\n{ec}\n\n"
        f"Failure evidence (for this probe):\n{fe}"
    )


def _judge_cell(jg: dict[str, Any]) -> str:
    if not jg:
        return ""
    err = jg.get("error")
    if err:
        return str(err)
    cs = jg.get("judgeConstraintSummary") if jg.get("judgeConstraintSummary") is not None else jg.get(
        "judge_constraint_summary",
    )
    brief_s = ""
    if isinstance(cs, dict) and cs:
        brief_s = _format_constraint_dict(cs).strip()
    v = jg.get("verdict")
    s = jg.get("score")
    top = ""
    if v is not None and str(v).strip():
        top = str(v).strip()
        if s is not None:
            top = f"{top} ({s})"
    elif s is not None:
        top = str(s)
    r = jg.get("reasoning")
    bottom = str(r).strip() if r else ""
    mid = f"{top}\n{bottom}" if top and bottom else (top or bottom)
    if brief_s and mid:
        return f"{brief_s}\n---\n{mid}"
    if brief_s:
        return brief_s
    return mid


def _merge_step_rows(
    total_steps: int,
    by_index: dict[int, dict[str, Any]],
    prompt_text_by_id: dict[str, str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for i in range(total_steps):
        live = by_index.get(i) or {}
        pid = live.get("promptId")
        prompt_text = prompt_text_by_id.get(pid) if pid else None
        ag = live.get("agent") or {}
        jg = live.get("judge") or {}
        rows.append(
            {
                "index": i,
                "prompt_text": prompt_text,
                "result": ag.get("responsePreview") if ag else None,
                "judge": _judge_cell(jg),
            },
        )
    return rows


def attack_test_run_to_xlsx_bytes(
    run: SessionAttackTestRun,
    prompt_text_by_id: dict[str, str],
) -> bytes:
    folded = fold_attack_test_events(run.events)
    step_rows = _merge_step_rows(folded.total_steps, folded.by_index, prompt_text_by_id)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Results"

    headers = ["#", "Prompt", "Result", "Judge"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h).font = Font(bold=True)

    for r, row in enumerate(step_rows, start=2):
        ws.cell(row=r, column=1, value=row["index"] + 1)
        ws.cell(row=r, column=2, value=row["prompt_text"] or "")
        ws.cell(row=r, column=3, value=row["result"] or "")
        ws.cell(row=r, column=4, value=row["judge"] or "")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def prompt_text_map_from_rows(prompts: list[Any] | None) -> dict[str, str]:
    if not prompts:
        return {}
    out: dict[str, str] = {}
    for p in prompts:
        out[str(p.id)] = p.prompt_text
    return out
